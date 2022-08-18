from cmath import nan
from django.views.decorators.csrf import csrf_exempt
from django.core.handlers.wsgi import WSGIRequest
from django.http import JsonResponse
from django.http import HttpResponse
import json
from HRCS.models import ShiftsTable
from HRCS.models import AttendanceTable
from HRCS.models import Token
from HRCS.models import EmpTable
from HRCS.models import JwtToken
from HRCS.models import AttendanceEmpView
from HRCS.models import ShiftsAttendanceEmpView
from HRCS.models import AttendanceEmpTodayView
from HRCS.models import AttendanceWorkOvertimeView
from HRCS.models import WorkOvertime
from HRCS.models import OvertimeEmpView
from HRCS.models import CompanyTable
from HRCS.models import GroupTable
from HRCS.models import DepTable
from HRCS.models import TimePeriod
from HRCS.models import WorkpositionList
from django.db.models import Q
from datetime import datetime as dt, timedelta
from functools import reduce
from openpyxl.utils import get_column_letter
import operator
import datetime
import jwt
import os
import numpy as np
from openpyxl.styles import Font, colors, Alignment   
import pandas as pd
import secrets
import calendar
from django.core.mail import send_mail
@csrf_exempt
def getDepdata(request: WSGIRequest):
    if request.method == 'GET':
        # 如果要用到其他資料表的，在這裡再多一個逗號對應不同的資料表就行
        return JsonResponse({'Response': _get_dep_data(request)})

@csrf_exempt
def verifyReset(request: WSGIRequest):
    if request.method == 'POST':
        # 如果要用到其他資料表的，在這裡再多一個逗號對應不同的資料表就行
        return JsonResponse({'Response': _verify_reset(request)})


@csrf_exempt
def getGroupdata(request: WSGIRequest):
    if request.method == 'GET':
        # 如果要用到其他資料表的，在這裡再多一個逗號對應不同的資料表就行
        return JsonResponse({'Response': _get_group_data(request)})
@csrf_exempt
def getWorkpositionList(request: WSGIRequest):
    if request.method == 'GET':
        # 如果要用到其他資料表的，在這裡再多一個逗號對應不同的資料表就行
        return JsonResponse({'Response': _get_WorkpositionList(request)})

@csrf_exempt
def getCompanydata(request: WSGIRequest):
    if request.method == 'GET':
        # 如果要用到其他資料表的，在這裡再多一個逗號對應不同的資料表就行
        return JsonResponse({'Response': _get_company_data(request)})

@csrf_exempt
def getWorkovertimeQuery(request: WSGIRequest):
    if request.method == 'GET':
        # 如果要用到其他資料表的，在這裡再多一個逗號對應不同的資料表就行
        return JsonResponse({'Response': _get_work_overtime_query(request)})

@csrf_exempt
def exportRMPQuery(request: WSGIRequest):
    if request.method == 'GET':
        # 如果要用到其他資料表的，在這裡再多一個逗號對應不同的資料表就行
        return _get_export_query(request)


@csrf_exempt
def getAttendanceEmpViewdata(request: WSGIRequest):
    if request.method == 'GET':
        return JsonResponse({'Response': _get_attendance_emp_view_data(request)})

@csrf_exempt
def getAttendanceEmpViewQuery(request: WSGIRequest):
    if request.method == 'GET':
        return JsonResponse({'Response': _get_attendance_emp_view_query(request)})
@csrf_exempt
def getShiftsAttendanceEmpViewQuery(request: WSGIRequest):
    if request.method == 'GET':
        return JsonResponse({'Response': _get_shifts_attendance_emp_view_query(request)})

@csrf_exempt
def Login(request: WSGIRequest):  # 登入測試
    if request.method == 'POST':
        return JsonResponse({'Response': _login(request)})
@csrf_exempt
def getEmpdata(request: WSGIRequest):
    if request.method == 'GET':
        # 如果要用到其他資料表的，在這裡再多一個逗號對應不同的資料表就行
        return JsonResponse({'Response': _get_emp_data(request)})
    if request.method == 'POST':
        return JsonResponse({'status': _insert_emp_data(request)})
    if request.method == 'PUT':
        return JsonResponse({'status': _update_emp_data(request)})

@csrf_exempt
def attendanceData(request: WSGIRequest):
    if request.method == 'POST':
        return JsonResponse({'status': _insert_attendance_data(request)})

@csrf_exempt
def getTimePeriod(request: WSGIRequest):
    if request.method == 'GET':
        # 如果要用到其他資料表的，在這裡再多一個逗號對應不同的資料表就行
        return JsonResponse({'Response': _get_time_period(request)})
    if request.method == 'POST':
        return JsonResponse({'status': _insert_time_period(request)})
    if request.method == 'PUT':
        return JsonResponse({'status': _update_time_period(request)})

@csrf_exempt
def RefreshJWTToken(request: WSGIRequest):  # 登入測試
    if request.method == 'POST':
        return JsonResponse({'Response': _refresh_jwt_token(request)})        
@csrf_exempt
def ForgetPassword(request: WSGIRequest):
    if request.method == 'POST':
        return JsonResponse({'status': _forget_password(request)})
    else:
        return JsonResponse({'status': 'error'})

@csrf_exempt
def ResetPassword(request: WSGIRequest):
    if request.method == 'POST':
        return JsonResponse({'status': _reset_password(request)})
    else:
        return JsonResponse({'status': 'error'})

@csrf_exempt
def Shiftdata(request: WSGIRequest):  # 出勤詳細資料表資料取得
    if request.method == 'GET':
        return JsonResponse({'Response': _get_shifts_data(request)})
    if request.method == 'POST':
        return JsonResponse({'status': _insert_shifts_data(request)})
    if request.method == 'PUT':
        return JsonResponse({'status': _update_Shift_data(request)})

@csrf_exempt
def Workovertime_del(request: WSGIRequest, overtime_no):  
    if request.method == 'DELETE':
        return JsonResponse({'status': _delete_Workovertime_data(request, overtime_no)})

@csrf_exempt
def Shiftdata_del(request: WSGIRequest, shiftno):  # 出勤詳細資料表資料刪除
    if request.method == 'DELETE':
        return JsonResponse({'status': _delete_Shift_data(request, shiftno)})

@csrf_exempt
def Delete_Attend_Punch(request: WSGIRequest):  # 登入測試
    if request.method == 'POST':
        return JsonResponse({'Response': _delete_attend_punch(request)})  

"""
@csrf_exempt
def getAttendancedata(request: WSGIRequest):
    if request.method == 'GET':
        # 如果要用到其他資料表的，在這裡再多一個逗號對應不同的資料表就行
        return JsonResponse({'Response': _get_attendance_data(request)})
"""

def _verify_jwt(token):
    verify=JwtToken.objects.filter(token=token)
    try:
        nowtime = dt.strptime(str((dt.utcnow()+datetime.timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')), '%Y-%m-%d %H:%M:%S')
        exp_time=dt.strptime(str(verify[0].exp_time.strftime('%Y-%m-%d %H:%M:%S')), '%Y-%m-%d %H:%M:%S')
        print('AAA')
        if nowtime-exp_time>timedelta(seconds=0):
            print('AAAA')
            raise Exception("over limit!")
        print(nowtime-exp_time)
    except Exception as e:
        print(e)
        return e
    

def _login(request):
    requestBody = request.body
    requestData = json.loads(requestBody)
    username =  requestData['username']
    password = requestData['password']
    errorMsg='error!'
    posts = []
    try:
        filter_post = EmpTable.objects.filter(Q(emp_no=username) & Q(user_pwd=password) &Q(working=1))
        payload = {
        'iss': 'CGPT',#改
        'sub': filter_post[0].emp_name,
        'aud': 'CGPT',#改
        'iat': dt.utcnow(),
        'username': filter_post[0].emp_name,
        'password':filter_post[0].user_pwd
        }
        try:
            token=jwt.encode(payload, 'secret', algorithm='HS256').decode('utf-8')#遠端
        except:
            token=jwt.encode(payload, 'secret', algorithm='HS256')

        #於雲端主機進行回傳時 必須轉換成字串 否則會觸發500error 
        #最大肇因是因為python3跟python2encode時的差異 但本地端python3不知道為啥跑不起來
        print(token)
        exp_time=(dt.utcnow()+datetime.timedelta(days=1)+datetime.timedelta(hours=8)).strftime("%Y-%m-%d %H:%M:%S")#一天有效時間&時區補正
        print(exp_time)
        created = JwtToken.objects.get_or_create(username=username)
        if created:#如果已經存在jwt_token
            _post =  JwtToken.objects.filter(username=username)
            _post.update(
            token=token,
            exp_time=exp_time
            )   
        else:#如果資料庫尚未存在此jwt_tokem
            JwtToken.objects.create(
                username=username,
                token=token,
                exp_time=exp_time
            )
        # person just refers to the existing one
        post_data={
                'emp_no':filter_post[0].emp_no,
                'emp_name':filter_post[0].emp_name,
                'auth':filter_post[0].auth,
                'group_no':filter_post[0].group_no,
                'dep_no':filter_post[0].dep_no,
                'group_leader':filter_post[0].group_leader,
                'token':token,
        }
        posts.append(post_data)

        return posts

    except Exception as e:
        if 'list index out of range' in str(e):
            errorMsg += '請重新確認輸入內容!'
        else:
            errorMsg += str(e)
        return errorMsg

def _get_shifts_data(request):  # 根據條件取得資料，若無條件則全取;回傳資料為JSON格式
    try:
        jwt_verify=_verify_jwt(request.headers['Token'])
        if "list index out of range" in str(jwt_verify):
            raise Exception("token不存在!")
        elif "over limit!" in str(jwt_verify):
            raise Exception("token已過期!")
    except Exception as e:
        Response = {
                'status': 'Error',
                'msg': str(e)
            }
        return Response
    
    emp_no = request.GET.get('emp_no')
    year = request.GET.get('year')
    month = request.GET.get('month')
    Response = {}
    posts = []
    filter_post = []
    status = ''

    # print('_get_shifts_data','attend_no',attend_no,'shift_no',shift_no,'emp_no',emp_no)

    #  根據有無額外資料選擇取得條件
    if year != None and month != None and emp_no != None:
        try:
            filter_post = ShiftsTable.objects.filter(
                Q(date__year=year) & Q(date__month=month) & Q(emp_no=emp_no))
            for _post in filter_post:
                post_data = {
                    'shift_no': _post.shift_no,
                    # 'attend_no': _post.attend_no,
                    'emp_no': _post.emp_no,
                    'date': _post.date,
                    'punch_in': _post.punch_in.strftime("%Y-%m-%d %H:%M:%S"),
                    'punch_out': _post.punch_out.strftime("%Y-%m-%d %H:%M:%S"),
                    'text': _post.text
                }

                posts.append(post_data)

            if posts == []:
                Response = {
                    'status': 'Null',
                    'data': posts
                }
            else:
                Response = {
                    'status': 'Succeed',
                    'data': posts
                }
            return Response
        except Exception as e:
            Response = {
                'status': 'Error',
                'msg': str(e)
            }
            return Response
    elif year != None and month != None:
        try:
            filter_post = ShiftsTable.objects.filter(
                Q(date__year=year) & Q(date__month=month))
            for _post in filter_post:
                post_data = {
                    'shift_no': _post.shift_no,
                    # 'attend_no': _post.attend_no,
                    'emp_no': _post.emp_no,
                    'date': _post.date,
                    'punch_in': _post.punch_in.strftime("%Y-%m-%d %H:%M:%S"),
                    'punch_out': _post.punch_out.strftime("%Y-%m-%d %H:%M:%S"),
                    'text': _post.text
                }

                posts.append(post_data)

            if posts == []:
                Response = {
                    'status': 'Null',
                    'data': posts
                }
            else:
                Response = {
                    'status': 'Succeed',
                    'data': posts
                }
            return Response
        except Exception as e:
            Response = {
                'status': 'Error',
                'msg': str(e)
            }
            return Response
    elif emp_no != None:
        try:
            filter_post = ShiftsTable.objects.filter(emp_no=emp_no)
            for _post in filter_post:
                post_data = {
                    'shift_no': _post.shift_no,
                    # 'attend_no': _post.attend_no,
                    'emp_no': _post.emp_no,
                    'date': _post.date,
                    'punch_in': _post.punch_in.strftime("%Y-%m-%d %H:%M:%S"),
                    'punch_out': _post.punch_out.strftime("%Y-%m-%d %H:%M:%S"),
                    'text': _post.text
                }

                posts.append(post_data)

            if posts == []:
                Response = {
                    'status': 'Null',
                    'data': posts
                }
            else:
                Response = {
                    'status': 'Succeed',
                    'data': posts
                }
            return Response
        except Exception as e:
            Response ={
                'status': 'Error',
                'msg': str(e)
            }
            return Response
    else:
        try:
            all_posts = ShiftsTable.objects.all()  # 取得所有資料
            for _post in all_posts:
                post_data = {
                    'shift_no': _post.shift_no,
                    # 'attend_no': _post.attend_no,
                    'emp_no': _post.emp_no,
                    'date': _post.date,
                    'punch_in': _post.punch_in.strftime("%Y-%m-%d %H:%M:%S"),
                    'punch_out': _post.punch_out.strftime("%Y-%m-%d %H:%M:%S"),
                    'text': _post.text
                }

                posts.append(post_data)

            if posts == []:
                Response = {
                    'status': 'Null',
                    'data': posts
                }
            else:
                Response = {
                    'status': 'Succeed',
                    'data': posts
                }
            return Response
        except Exception as e:
            Response = {
                'status': 'Error',
                'msg': str(e)
            }
            return Response
def _get_group_data(request):
    try:
        jwt_verify=_verify_jwt(request.headers['Token'])
        if "list index out of range" in str(jwt_verify):
            raise Exception("token不存在!")
        elif "over limit!" in str(jwt_verify):
            raise Exception("token已過期!")
        all_posts = GroupTable  .objects.all()
        posts = []

        for _post in all_posts:  # 按照api格式塞進去
            post_data = {  # 按照api格式塞進去
            'group_no':_post.group_no,
            'group_name':_post.group_name,
             }
            posts.append(post_data)

            if posts == []:
                Response = {
                    'status': 'Null',
                    'data': posts
                }
            else:
                Response = {
                    'status': 'Succeed',
                    'data': posts
                }
            #return Response
    except Exception as e:
        Response = {
            'status': 'Error',
            'msg': str(e)
        }

    
    return Response
def _get_company_data(request):
    try:
        jwt_verify=_verify_jwt(request.headers['Token'])
        if "list index out of range" in str(jwt_verify):
            raise Exception("token不存在!")
        elif "over limit!" in str(jwt_verify):
            raise Exception("token已過期!")
        all_posts = CompanyTable .objects.all()

        posts = []

        for _post in all_posts:  # 按照api格式塞進去
            post_data = {  # 按照api格式塞進去
            'comp_no':_post.comp_no,
            'comp_name':_post.comp_name,
            'main':_post.main
            }
            posts.append(post_data)

            if posts == []:
                Response = {
                    'status': 'Null',
                    'data': posts
                }
            else:
                Response = {
                    'status': 'Succeed',
                    'data': posts
                }
            #return Response
    except Exception as e:
        Response = {
            'status': 'Error',
            'msg': str(e)
        }

    
    return Response

def _get_dep_data(request):
    try:
        jwt_verify=_verify_jwt(request.headers['Token'])
        if "list index out of range" in str(jwt_verify):
            raise Exception("token不存在!")
        elif "over limit!" in str(jwt_verify):
            raise Exception("token已過期!")
        all_posts = DepTable.objects.all()
        posts = []

        for _post in all_posts:  # 按照api格式塞進去
            post_data = {  # 按照api格式塞進去
            'dep_no':_post.dep_no,
            'dep_name':_post.dep_name,
             }
            posts.append(post_data)

            if posts == []:
                Response = {
                    'status': 'Null',
                    'data': posts
                }
            else:
                Response = {
                    'status': 'Succeed',
                    'data': posts
                }
            #return Response
    except Exception as e:
        Response = {
            'status': 'Error',
            'msg': str(e)
        }

    
    return Response



    
def _insert_shifts_data(request):  # 新增班表
    try:
        jwt_verify=_verify_jwt(request.headers['Token'])
        if "list index out of range" in str(jwt_verify):
            raise Exception("token不存在!")
        elif "over limit!" in str(jwt_verify):
            raise Exception("token已過期!")
    except Exception as e:
        Response = {
                'status': 'Error',
                'msg': str(e)
            }
        return Response

    requestBody = request.body
    requestData = json.loads(requestBody)
    errorMsg = '發生錯誤!\n'
    today = datetime.date.today()
    try:  
        if(requestData['del']):
            if int(requestData['date'].split("-")[1]) > datetime.date(today.year, today.month, calendar.monthrange(today.year, today.month)[1]).month:        
                ShiftsTable.objects.filter(
                    Q(emp_no=requestData['emp_no'])& 
                    Q(date__gte = datetime.datetime(int(requestData['date'].split("-")[0]), int(requestData['date'].split("-")[1]), int(requestData['date'].split("-")[2])))& 
                    Q(date__lte =datetime.datetime(int(requestData['date'].split("-")[0]), int(requestData['date'].split("-")[1]), calendar.monthrange(int(requestData['date'].split("-")[0]), int(requestData['date'].split("-")[1]))[1]))).delete()     
            else:
                filter_list=[Q(emp_no=requestData['emp_no'])]
                if(int(requestData['date'].split("-")[2])<=today.day):
                    filter_list.append(Q(date__gt = today)) 
                else:
                    filter_list.append(Q(date__gte = datetime.datetime(int(requestData['date'].split("-")[0]), int(requestData['date'].split("-")[1]), int(requestData['date'].split("-")[2]))))          
                
                filter_list.append(Q(date__lte =datetime.datetime(today.year, today.month, calendar.monthrange(today.year, today.month)[1])))
                ShiftsTable.objects.filter(reduce(operator.and_, filter_list)).delete()
      
            _posts = ShiftsTable.objects.filter(Q(shift_no=requestData['shift_no']))
            if len(_posts) <= 0:   
                ShiftsTable.objects.create(
                    shift_no=requestData['shift_no'],
                    emp_no=requestData['emp_no'],
                    date=requestData['date'],
                    punch_in=requestData['punch_in'],
                    punch_out=requestData['punch_out'],
                    text=requestData['text']
                )

    except Exception as e:
        if 'Duplicate entry' in str(e):
            errorMsg += requestData['shift_no']+'此編號已有資料，請重新確認!'
        else:
            errorMsg += str(e)
        return errorMsg

    return 'Succeed'
    
def _refresh_jwt_token(request):  # 更新jwttoken
    requestBody = request.body
    requestData = json.loads(requestBody)
    errorMsg = '無法更新token有效時間!'

    try:
        print("in")
        _post = JwtToken.objects.filter(token=requestData['token'])
        exp_time= dt.strptime(_post[0].exp_time.strftime('%Y-%m-%d %H:%M:%S'),'%Y-%m-%d %H:%M:%S')
        nowtime = dt.strptime(str((dt.utcnow()+datetime.timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')), '%Y-%m-%d %H:%M:%S')
        print(exp_time)
        print(nowtime)
        if nowtime-exp_time<timedelta(seconds=0):#現在時間減去過期時間 大於0代表已過期
            exp_time=(dt.utcnow()+datetime.timedelta(days=1)+datetime.timedelta(hours=8)).strftime("%Y-%m-%d %H:%M:%S")#一天有效時間&時區補正
            _post.update(
                exp_time=exp_time
            )
        else:
            raise Exception("over limit!")

    except Exception as e:
        if 'over limit!' in str(e):
            errorMsg+="token已過期!"
        elif 'list index out of range' in str(e):
            errorMsg+="查無此token!"
        else:
            errorMsg += str(e)
        return errorMsg

    return 'Succeed'

def _update_Shift_data(request):  # 更新出勤資料
    try:
        jwt_verify=_verify_jwt(request.headers['Token'])
        if "list index out of range" in str(jwt_verify):
            raise Exception("token不存在!")
        elif "over limit!" in str(jwt_verify):
            raise Exception("token已過期!")
    except Exception as e:
        Response = {
                'status': 'Error',
                'msg': str(e)
            }
        return Response

    requestBody = request.body
    requestData = json.loads(requestBody)
    errorMsg = '修改錯誤!'

    try:
        print( (requestData['shift_no'][0:-8])+(requestData['date'].replace("-","")))
        _post = ShiftsTable.objects.filter(shift_no=requestData['shift_no'])
        _post.update(
            shift_no= (requestData['shift_no'][0:-8])+(requestData['date'].replace("-","")),
            date=requestData['date'],
            punch_in=requestData['punch_in'],
            punch_out=requestData['punch_out'],
            text=requestData['text']
        )



    except Exception as e:
        return errorMsg

    return 'Succeed'

def _delete_Shift_data(request, shiftno):
    try:
        jwt_verify=_verify_jwt(request.headers['Token'])
        if "list index out of range" in str(jwt_verify):
            raise Exception("token不存在!")
        elif "over limit!" in str(jwt_verify):
            raise Exception("token已過期!")
    except Exception as e:
        Response = {
                'status': 'Error',
                'msg': str(e)
            }
        return Response

    errorMsg = '刪除錯誤!'

    try:
        ShiftsTable.objects.filter(shift_no=shiftno).delete()
        # _post = ShiftsTable.objects.filter(shift_no=shiftno)
        # print(_post)
    except Exception as e:
        return errorMsg

    return 'Succeed'

def _delete_Workovertime_data(request, overtime_no):
    try:
        jwt_verify=_verify_jwt(request.headers['Token'])
        if "list index out of range" in str(jwt_verify):
            raise Exception("token不存在!")
        elif "over limit!" in str(jwt_verify):
            raise Exception("token已過期!")
    except Exception as e:
        Response = {
                'status': 'Error',
                'msg': str(e)
            }
        return Response

    errorMsg = '刪除錯誤!'

    try:
        WorkOvertime.objects.filter(overtime_no=overtime_no).delete()
        # _post = ShiftsTable.objects.filter(shift_no=shiftno)
        # print(_post)
    except Exception as e:
        return errorMsg

    return 'Succeed'

def _verify_reset(request):

    requestBody = request.body
    requestData = json.loads(requestBody)
    errorMsg = '驗證失敗!'
    # print("in")
    try:
        _token= Token.objects.filter(token=requestData['token'])#從token資料表取出該名員工的token
        nowtime = dt.strptime(str((dt.utcnow()+datetime.timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')), '%Y-%m-%d %H:%M:%S')
        deadline=dt.strptime(str(_token[0].deadline.strftime('%Y-%m-%d %H:%M:%S')), '%Y-%m-%d %H:%M:%S')
        if nowtime-deadline<timedelta(seconds=0):
            Response = {
                'status': 'Succeed',
                'emp_no': _token[0].emp_no
            }
        else:
            errorMsg += '驗證碼已過期!'
            Response = {
                'status': 'Error',
                'msg': errorMsg
            }

    except Exception as e:
        if 'list index out of range' in str(e):
            errorMsg += '驗證碼不存在!'
        else:
            errorMsg += str(e)
        Response = {
            'status': 'Error',
            'msg': errorMsg
        }
    return Response

def _reset_password(request):  # 忘記密碼重設請求_驗證

    requestBody = request.body
    requestData = json.loads(requestBody)
    errorMsg = '修改失敗!'
    # print("in")
    try:
        _post = EmpTable.objects.filter(emp_no=requestData['emp_no'])
        _token= Token.objects.filter(emp_no=requestData['emp_no'])#從token資料表取出該名員工的token
        nowtime = dt.strptime(str((dt.utcnow()+datetime.timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')), '%Y-%m-%d %H:%M:%S')
        deadline=dt.strptime(str(_token[0].deadline.strftime('%Y-%m-%d %H:%M:%S')), '%Y-%m-%d %H:%M:%S')
        if _token[0].token==requestData['token'] and nowtime-deadline<timedelta(seconds=0):#token正確且時間在deadline前
            print("yes")
            _post.update(user_pwd=requestData['password'])  # 更新密碼
            _token.delete()#刪除重設密碼的token
            return 'Succeed'
        elif _token[0].token!=requestData['token']:
            return '驗證碼錯誤!'
        elif nowtime-deadline>timedelta(seconds=0):
            return '驗證碼已過期'

        #_post.update(user_pwd=requestData['password'])  # UPDATE(ry
    except Exception as e:
        if 'list index out of range' in str(e):
            errorMsg += '查無此帳號!請重新確認輸入內容!'
        else:
            errorMsg += str(e)
        return errorMsg

def _forget_password(request):#忘記密碼驗證&寄出驗證信
    requestBody = request.body
    requestData = json.loads(requestBody)
    errorMsg = '寄送失敗!'
    # print("in")
    try:
        _post = EmpTable.objects.filter(Q(emp_no=requestData['emp_no']) & Q(mail=requestData['mail']))
        email=_post[0].mail#確保是使用者的信箱
        deadline=requestData['deadline']
        token=secrets.token_urlsafe(20)#產生亂數token
        url=requestData['url']
        created = Token.objects.get_or_create(emp_no=requestData['emp_no'])
        if created:#如果已經存在jwt_token
            _post =  Token.objects.filter(emp_no=requestData['emp_no'])
            _post.update(
            token=token,
            deadline=deadline
            )   
        else:#如果資料庫尚未存在此jwt_tokem
            Token.objects.create(#紀錄重置密碼的token
            emp_no=requestData['emp_no'],
            token=token,
            deadline=deadline
            )

        print(token)
        send_mail('密碼重設信件', '您好，請透過以下連結進行密碼重設的動作:\n'+url+token+'\n如果無法點擊連結，請將網址複製貼上至瀏覽器即可。 ', 'cgptiot@gmail.com', [email])
        #標題.內容.寄信者的信箱.收信者的信箱


        print("發送成功!")
    except Exception as e:
        if 'list index out of range' in str(e):
            errorMsg += '帳號或信箱輸入錯誤!請重新確認輸入內容!'
        else:
            errorMsg += str(e)
        return errorMsg

    return 'Succeed'

def _get_export_query(request):
    path =  os.getcwd()+r"\\emp_Report\\"
    if not os.path.isdir(path):
        os.mkdir(path)

    emp_no = request.GET.get('emp_no')

    if(request.GET.get('date_start')==None):
        now = datetime.datetime.now().date()
        this_month_start = datetime.datetime(now.year, now.month, 1)
        this_month_end = datetime.datetime(now.year, now.month, calendar.monthrange(now.year, now.month)[1])
    else:
        this_month_start=dt.strptime(request.GET.get('date_start'),'%Y-%m-%d')
        this_month_end=dt.strptime(request.GET.get('date_end'),'%Y-%m-%d')

    all_posts = AttendanceWorkOvertimeView.objects.filter(Q(emp_no=emp_no)&Q(attend_date__range=(this_month_start,this_month_end))).order_by('attend_date')  # 自模型取出資料
    
    posts=[]
    month_every =this_month_start
    i=0
    if len(all_posts) > 0:
        while month_every <= this_month_end:
            if(month_every.date() == all_posts[i].attend_date):
                punch_in=None if all_posts[i].punch_in==None else all_posts[i].punch_in.strftime("%H:%M:%S")#三元運算子
                punch_out=None if all_posts[i].punch_out==None else all_posts[i].punch_out.strftime("%H:%M:%S")
                st_punch_in=None if all_posts[i].st_punch_in==None else all_posts[i].st_punch_in.strftime("%H:%M:%S")#三元運算子
                st_punch_out=None if all_posts[i].st_punch_out==None else all_posts[i].st_punch_out.strftime("%H:%M:%S")
                overtime_punch_in=None if all_posts[i].overtime_punch_in==None else all_posts[i].overtime_punch_in.strftime("%H:%M:%S")#三元運算子         
                overtime_punch_out=None if all_posts[i].overtime_punch_out==None else all_posts[i].overtime_punch_out.strftime("%H:%M:%S")
                
                overtime_hour=None
                if(all_posts[i].start_time!=None and all_posts[i].end_time!=None):
                    overtime_sec=datetime.datetime.strptime(all_posts[i].end_time.strftime("%H:%M:%S"),"%H:%M:%S")-datetime.datetime.strptime(all_posts[i].start_time.strftime("%H:%M:%S"),"%H:%M:%S")
                    overtime_hour=overtime_sec.seconds/3600

                    overtime_check=None
                    if all_posts[i].overtime_hours != None:
                        if(int(overtime_sec.seconds/60) > all_posts[i].overtime_hours):
                            overtime_check="未達時數"

               
                post_data = {  # 按照api格式塞進去d
                    # '員工編號':all_posts[i].emp_no, 
                    # '員工姓名':all_posts[i].emp_name,
                    # '職位':all_posts[i].wp_name,
                    # '組別':all_posts[i].group_name,
                    '上班日期':all_posts[i].attend_date,
                    '上班打卡時間':punch_in,
                    '排班(上)':st_punch_in,
                    '排班(下)':st_punch_out,
                    '下班打卡時間':punch_out,
                    # '上班打卡地點':all_posts[i].in_position,
                    # '下班打卡地點':all_posts[i].out_position,
                    '今日工時(小時)':all_posts[i].total_time,
                    '遲到時間(分鐘)':all_posts[i].latetime,
                    '加班打卡時間':overtime_punch_in,
                    '加班結束打卡時間':overtime_punch_out,
                    '加班時數檢查':overtime_check,
                    '加班工時(小時)':overtime_hour
                }

                if(i < len(all_posts)-1):
                    i=i+1
                    
                posts.append(post_data)
            else:
                post_data = {  # 按照api格式塞進去
                    # '員工編號':all_posts[0].emp_no, 
                    # '員工姓名':all_posts[0].emp_name,
                    # '職位':all_posts[0].wp_name,
                    # '組別':all_posts[0].group_name,
                    '上班日期':month_every.date(),
                    '上班打卡時間':nan,
                    '排班(上)':nan,                   
                    '下班打卡時間':nan,
                    '排班(下)':nan,
                    # '上班打卡地點':None,
                    # '下班打卡地點':None,
                    '今日工時(小時)':nan,
                    '遲到時間(分鐘)':nan,
                    '加班打卡時間':nan,
                    '加班結束打卡時間':nan,
                    '加班時數檢查':nan,
                    '加班工時(小時)':nan
                }

                posts.append(post_data)

            month_every+= timedelta(days=1)

        df = pd.DataFrame(posts)


        
        totaltime=0
        totalLatetime=0
        totalOvertime=0
        for x in df.index:
            #工時處理
            if(df.loc[x, "今日工時(小時)"]) >540:
                df.loc[x, "今日工時(小時)"] =540

            if pd.notna(df.loc[x, "今日工時(小時)"]) :        
                df.loc[x, "今日工時(小時)"] = round((df.loc[x, "今日工時(小時)"]/60),2)-round((df.loc[x, "今日工時(小時)"]/60),2)%0.5
                totaltime=  totaltime+df.loc[x, "今日工時(小時)"]

            if  pd.notna(df.loc[x, "遲到時間(分鐘)"]):
                totalLatetime= totalLatetime +df.loc[x, "遲到時間(分鐘)"]

            if   pd.notna(df.loc[x, "加班工時(小時)"]):
                totalOvertime= totalOvertime+df.loc[x, "加班工時(小時)"]

            if pd.notna(df.loc[x, "排班(上)"]) & pd.isna(df.loc[x, "上班打卡時間"]): 
                df.loc[x, "上班打卡時間"] ="未打卡"

            if pd.notna(df.loc[x, "排班(下)"]) & pd.isna(df.loc[x, "下班打卡時間"]): 
                df.loc[x, "下班打卡時間"] ="未打卡"

        #加總
        df.loc[len(df.index), "上班日期"] = "總計"
        df.loc[len(df.index)-1, "今日工時(小時)"] = totaltime
        df.loc[len(df.index)-1, "遲到時間(分鐘)"] = totalLatetime
        df.loc[len(df.index)-1, "加班工時(小時)"] = totalOvertime
        
        FilePath = os.getcwd()+r"\\emp_Report\\"+all_posts[0].emp_name +".xlsx"  

        writer = pd.ExcelWriter(FilePath, engine = 'openpyxl',)
        df.to_excel(writer, startrow = 1, sheet_name = 'Sheet1', index=False)
        #Excel自動調整寬度
        column_widths = (
            df.columns.to_series().apply(lambda x: len(str(x).encode('gbk'))).values
        )
        max_widths = (
            df.astype(str).applymap(lambda x: len(str(x).encode('gbk'))).agg(max).values
        )
        widths = np.max([column_widths+3, max_widths+3], axis=0)
        worksheet = writer.sheets['Sheet1']

        for i, width in enumerate(widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width + 2
            
        # 调整列宽
        # worksheet.column_dimensions['G'].width = 100.0
        # 调整行高
        worksheet.row_dimensions[1].height = 25
        #合併
        worksheet.merge_cells('A1:J1')

        font =Font(size=18)
        worksheet.cell(1, 1, 
        "員工編號："+all_posts[0].emp_no+
        "  員工姓名："+all_posts[0].emp_name+
        "  職稱："+all_posts[0].wp_name+
        "  組別："+all_posts[0].group_name).font = font
        
        for i in range(1, worksheet.max_row+1):
            for j in range(1, worksheet.max_column+1):
                worksheet.cell(row = i, column = j).alignment = Alignment(horizontal='center', vertical='center')
      
        
        #writer.save()
        writer.close()

        with open(FilePath, 'rb') as model_excel:
            result = model_excel.read()
            response = HttpResponse(result)
            response['Content-Type'] = 'application/octet-stream' 
            response['Content-Disposition'] = 'attachment; filename=test.xlsx'
            return response    
    else:
        response = HttpResponse()
        response.status_code=500
        return response
    

def _get_work_overtime_query(request):
    try:
        jwt_verify=_verify_jwt(request.headers['Token'])
        if "list index out of range" in str(jwt_verify):
            raise Exception("token不存在!")
        elif "over limit!" in str(jwt_verify):
            raise Exception("token已過期!")
    except Exception as e:
        Response = {
                'status': 'Error',
                'msg': str(e)
            }
        return Response
    emp_no = request.GET.get('emp_no')
    start_date= request.GET.get('start_date')
    end_date=request.GET.get('end_date')
    dep_no=request.GET.get('dep_no')
    work_position=request.GET.get('work_position')
    group_no=request.GET.get('group_no')
    if start_date!=None and end_date!=None:
        #try:
        filter_list=[Q(overtime_date__range=(start_date,end_date))]
        if emp_no!=None and emp_no!="":
            filter_list.append(Q(emp_no=emp_no))
        if dep_no!=None and dep_no!="":
            filter_list.append(Q(dep_no=dep_no))
        if work_position!=None and work_position!="":
            filter_list.append(Q(work_position=work_position))
        if group_no!=None and group_no!="":
            filter_list.append(Q(group_no=group_no))          
        
        all_posts = OvertimeEmpView.objects.filter(reduce(operator.and_, filter_list))  # 自模型取出資料
        print(all_posts)
        posts = []
        for _post in all_posts:  # 按照api格式塞進去
            punch_in="Null" if _post.punch_in==None else _post.punch_in.strftime("%Y-%m-%d %H:%M:%S")#三元運算子
            punch_out="Null" if _post.punch_out==None else _post.punch_out.strftime("%Y-%m-%d %H:%M:%S")
            post_data = {  # 按照api格式塞進去
                'overtime_id':_post.overtime_id, 
                'overtime_no':_post.overtime_no,
                'attend_no':_post.attend_no,
                'overtime_pay_no':_post.overtime_pay_no,
                'overtime_date':_post.overtime_date,
                'emp_no':_post.emp_no,
                'punch_in':punch_in,
                'punch_out':punch_out,
                'overtime_hours':_post.overtime_hours,
                'overtime_check':_post.overtime_check,
                'dep_no':_post.dep_no,
                'work_position':_post.work_position,
                'emp_name':_post.emp_name,
                'group_no':_post.group_no
                #'overtime_type_no':_post.overtime_type_no,
                #'hours_change':_post.hours_change,
                #'disaster':_post.disaster,
                #'no_rest':_post.no_rest
            }
            posts.append(post_data)

        if posts == []:
            Response = {
                    'status': 'Null',
                    'data': posts
            }
        else:
            Response = {
                    'status': 'Succeed',
                    'data': posts
            }
        return Response
    else:
        Response = {
            'status': 'Succeed',
            'msg': '查無資料'
        }
        return Response


def _get_attendance_emp_view_query(request):  # 取得出勤管理 基礎頁面的資料 需求emp_no,year,month
    try:
        jwt_verify=_verify_jwt(request.headers['Token'])
        if "list index out of range" in str(jwt_verify):
            raise Exception("token不存在!")
        elif "over limit!" in str(jwt_verify):
            raise Exception("token已過期!")
    except Exception as e:
        Response = {
                'status': 'Error',
                'msg': str(e)
            }
        return Response
    status = ''
    Response = {}
    emp_no = request.GET.get('emp_no')
    group_no=request.GET.get('group_no')
    start_date= request.GET.get('start_date')
    end_date=request.GET.get('end_date')
    if start_date!=None and end_date!=None:
        print(start_date)
        try:
            filter_list=[Q(attend_date__range=(start_date,end_date))]
            if emp_no!=None and emp_no!="":
                filter_list.append(Q(emp_no=emp_no))
            if group_no!=None and group_no!="":
                filter_list.append(Q(group_no=group_no))
            all_posts = AttendanceEmpView.objects.filter(reduce(operator.and_, filter_list))  # 自模型取出資料
            posts = []
            for _post in all_posts:  # 按照api格式塞進去
                punch_in=None if _post.punch_in==None else _post.punch_in.strftime("%Y-%m-%d %H:%M:%S")#三元運算子
                punch_out=None if _post.punch_out==None else _post.punch_out.strftime("%Y-%m-%d %H:%M:%S")
                #if _post.group_no!=0 and _post.group_no!="0":
                post_data = {  # 按照api格式塞進去
                        'attend_no':_post.attend_no,
                        'attend_date':_post.attend_date,
                        'comp_name':_post.comp_name,
                        'shift_no':_post.shift_no,
                        'emp_no': _post.emp_no,
                        'emp_name':_post.emp_name,
                        'group_no':_post.group_no,
                        'punch_in':punch_in,
                        'punch_out':punch_out,
                        'in_position':_post.in_position,
                        'out_position':_post.out_position,
                        'overtime_no':_post.overtime_no,
                        'salary_no':_post.salary_no,
                        'total_time':_post.total_time,
                        'fulltime':_post.fulltime,
                        'latetime':_post.latetime,
                        'attend_id':_post.attend_id,
                        'tel':_post.tel,
                        'work_position':_post.work_position,
                        'in_addr':_post.in_addr,
                        'out_addr':_post.out_addr
                    }
                posts.append(post_data)

            if posts == []:
                Response = {
                    'status': 'Null',
                    'data': posts
                }
            else:
                Response = {
                    'status': 'Succeed',
                    'data': posts
                }
            
        except Exception as e:
            Response = {
                    'status': 'Error',
                    'msg': str(e)
            }
    else:
        Response = {
            'status': 'Error',
            'msg': '缺少日期範圍!'
        }
    
    return Response


def _get_shifts_attendance_emp_view_query(request):  # 取得出勤管理 基礎頁面的資料 需求emp_no,year,month
    try:
        jwt_verify=_verify_jwt(request.headers['Token'])
        if "list index out of range" in str(jwt_verify):
            raise Exception("token不存在!")
        elif "over limit!" in str(jwt_verify):
            raise Exception("token已過期!")
    except Exception as e:
        Response = {
                'status': 'Error',
                'msg': str(e)
            }
        return Response
    status = ''
    Response = {}
    emp_no = request.GET.get('emp_no')
    group_no=request.GET.get('group_no')
    start_date= request.GET.get('start_date')
    end_date=request.GET.get('end_date')
    if start_date!=None and end_date!=None:
        try:
            filter_list=[Q(attend_date__range=(start_date,end_date))]
            if emp_no!=None and emp_no!="":
                filter_list.append(Q(emp_no=emp_no))
            if group_no!=None and group_no!="":
                filter_list.append(Q(group_no=group_no))
            all_posts = ShiftsAttendanceEmpView.objects.filter(reduce(operator.and_, filter_list))  # 自模型取出資料
            posts = []
            for _post in all_posts:  # 按照api格式塞進去
                punch_in=None if _post.punch_in==None else _post.punch_in.strftime("%Y-%m-%d %H:%M:%S")#三元運算子
                punch_out=None if _post.punch_out==None else _post.punch_out.strftime("%Y-%m-%d %H:%M:%S")
                s_punch_in=None if _post.s_punch_in==None else _post.s_punch_in.strftime("%Y-%m-%d %H:%M:%S")#三元運算子
                s_punch_out=None if _post.s_punch_out==None else _post.s_punch_out.strftime("%Y-%m-%d %H:%M:%S")
                #if _post.group_no!=0 and _post.group_no!="0":
                post_data = {  # 按照api格式塞進去
                        'attend_no':_post.attend_no,
                        'attend_date':_post.attend_date,
                        'comp_name':_post.comp_name,
                        'shift_no':_post.shift_no,
                        'emp_no': _post.emp_no,
                        'emp_name':_post.emp_name,
                        'group_no':_post.group_no,
                        'punch_in':punch_in,
                        'punch_out':punch_out,
                        's_punch_in':s_punch_in,
                        's_punch_out':s_punch_out,
                        'in_position':_post.in_position,
                        'out_position':_post.out_position,
                        'overtime_no':_post.overtime_no,
                        'salary_no':_post.salary_no,
                        'total_time':_post.total_time,
                        'fulltime':_post.fulltime,
                        'latetime':_post.latetime,
                        'attend_id':_post.attend_id,
                        'tel':_post.tel,
                        'work_position':_post.work_position,
                        'in_addr':_post.in_addr,
                        'out_addr':_post.out_addr
                    }
                posts.append(post_data)

            if posts == []:
                Response = {
                    'status': 'Null',
                    'data': posts
                }
            else:
                Response = {
                    'status': 'Succeed',
                    'data': posts
                }
            
        except Exception as e:
            Response = {
                    'status': 'Error',
                    'msg': str(e)
            }
    else:
        Response = {
            'status': 'Error',
            'msg': '缺少日期範圍!'
        }
    
    return Response

def _insert_attendance_data(request):
    try:
        jwt_verify=_verify_jwt(request.headers['Token'])
        if "list index out of range" in str(jwt_verify):
            raise Exception("token不存在!")
        elif "over limit!" in str(jwt_verify):
            raise Exception("token已過期!")
    except Exception as e:
        Response=str(e)
        return Response
    requestBody = request.body
    requestData = json.loads(requestBody)
    errorMsg = 'ERROR!'

    try:
        _post=AttendanceTable.objects.filter(Q(emp_no=requestData['emp_no'])& Q(attend_date=requestData['attend_date']))
        _post_shift=ShiftsTable.objects.filter(Q(emp_no=requestData['emp_no'])& Q(date=requestData['attend_date']))

        punch_in=None if _post_shift[0].punch_in==None else _post_shift[0].punch_in.strftime("%Y-%m-%d %H:%M:%S")
        punch_out=None if _post_shift[0].punch_out==None else _post_shift[0].punch_out.strftime("%Y-%m-%d %H:%M:%S")
         
        if(_post.exists() == 0):
            newAttendancedata = AttendanceTable.objects.create(
                attend_id = None,
                attend_date=requestData['attend_date'],
                attend_no="ATT"+requestData['emp_no']+requestData['attend_date'].replace("-",""),
                shift_no="SH"+requestData['emp_no']+requestData['attend_date'].replace("-",""),
                salary_no="SN"+requestData['emp_no']+requestData['attend_date'].replace("-","")[0:-2],
                in_position="25.081883122848833,121.59307408268548",
                latetime=0,
                punch_in =punch_in,
                emp_no=requestData['emp_no'],
            )

            _post_shift.update(
                attend_no ="ATT"+requestData['emp_no']+requestData['attend_date'].replace("-","")
            )
            
        if(requestData['time']):
            timedel = (_post_shift[0].punch_out) - (_post[0].punch_in)              
            AttendanceTable.objects.filter(Q(emp_no=requestData['emp_no'])& Q(attend_date=requestData['attend_date'])).update(
                out_position="25.081883122848833,121.59307408268548",
                punch_out =punch_out,
                total_time=(timedel.seconds)/60,
                fulltime=(timedel.seconds)/60/540,
            )

    except Exception as e:
        if 'list index out of range' in str(e):
            errorMsg += '此員工不再當日排班表中'
        else:
            errorMsg += str(e)
        return errorMsg

    return 'Succeed'


def _insert_emp_data(request):
    try:
        jwt_verify=_verify_jwt(request.headers['Token'])
        if "list index out of range" in str(jwt_verify):
            raise Exception("token不存在!")
        elif "over limit!" in str(jwt_verify):
            raise Exception("token已過期!")
    except Exception as e:
        Response=str(e)
        return Response
    requestBody = request.body
    requestData = json.loads(requestBody)
    errorMsg = 'ERROR!'

    try:
        newEmpdata = EmpTable.objects.create(
            emp_id = None,
            emp_no=requestData['emp_no'],
            emp_name=requestData['emp_name'],
            comp_no=requestData['comp_no'],
            dep_no=requestData['dep_no'],
            work_position=requestData['work_position'],
            user_pwd=requestData['user_pwd'],
            tel=requestData['tel'],
            group_no=requestData['group_no'],
            group_leader=requestData['group_leader'],
            working=requestData['working'],
            auth=requestData['auth'],
            mail=requestData['mail']

        )
    except Exception as e:
        if 'Duplicate entry' in str(e):
            errorMsg += '該編號重複，請再確認編號'
        else:
            errorMsg += str(e)
        return errorMsg

    return 'Succeed'

def _update_emp_data(request):
    try:
        jwt_verify=_verify_jwt(request.headers['Token'])
        if "list index out of range" in str(jwt_verify):
            raise Exception("token不存在!")
        elif "over limit!" in str(jwt_verify):
            raise Exception("token已過期!")
    except Exception as e:
        Response=str(e)
        return Response
    requestBody = request.body
    requestData = json.loads(requestBody)
    errorMsg = '修改錯誤!'
    try:
        _post = EmpTable.objects.filter(emp_no=requestData['emp_no'])
        _post.update(
            working=requestData['working'],
            emp_name=requestData['emp_name'],
            comp_no=requestData['comp_no'],
            dep_no=requestData['dep_no'],
            work_position=requestData['work_position'],
            tel=requestData['tel'],
            group_no=requestData['group_no'],
            group_leader=requestData['group_leader'],
            auth=requestData['auth'],
            mail=requestData['mail']
            )
    except Exception as e:
        errorMsg += str(e)
        return errorMsg
    return 'Succeed'

def _get_emp_data(request):  # 員工
    try:
        jwt_verify=_verify_jwt(request.headers['Token'])
        if "list index out of range" in str(jwt_verify):
            raise Exception("token不存在!")
        elif "over limit!" in str(jwt_verify):
            raise Exception("token已過期!")
        all_posts = EmpTable.objects.all()
        posts = []

        for _post in all_posts:  # 按照api格式塞進去
            post_data = {  # 按照api格式塞進去
            'emp_no':_post.emp_no,
            'emp_name':_post.emp_name,
            'comp_no':_post.comp_no,
            'dep_no':_post.dep_no,
            'work_position':_post.work_position,
            'tel':_post.tel,
            'group_no':_post.group_no,
            'group_leader':_post.group_leader,
            'working':_post.working,
            'auth':_post.auth,
            'mail':_post.mail
            }
            posts.append(post_data)

        if posts == []:
            Response = {
                'status': 'Null',
                'data': posts
            }
        else:
            Response = {
                'status': 'Succeed',
                'data': posts
            }
        return Response
    except Exception as e:
        Response = {
            'status': 'Error',
            'msg': str(e)
        }
    return Response

def _get_attendance_emp_view_data(request):  # 取得出勤管理 基礎頁面的資料 需求emp_no,year,month
    try:
        jwt_verify=_verify_jwt(request.headers['Token'])
        if "list index out of range" in str(jwt_verify):
            raise Exception("token不存在!")
        elif "over limit!" in str(jwt_verify):
            raise Exception("token已過期!")
    except Exception as e:
        Response = {
                'status': 'Error',
                'msg': str(e)
            }
        return Response
    status = ''
    Response = {}
    emp_no = request.GET.get('emp_no')
    year = request.GET.get('year')
    month = request.GET.get('month')
    day=request.GET.get('day')

    if emp_no!= None and year!=None and month!=None and day!=None: 
        try:
            all_posts = AttendanceEmpView.objects.filter(Q(emp_no=request.GET.get('emp_no')) & Q(working=1) &
            Q(attend_date__year=request.GET.get('year')) & Q(attend_date__month=request.GET.get('month')) & Q(attend_date__day=request.GET.get('day'))
            )  # 自模型取出資料
            posts = []

            for _post in all_posts:  # 按照api格式塞進去
                punch_in=None if _post.punch_in==None else _post.punch_in.strftime("%Y-%m-%d %H:%M:%S")#三元運算子
                punch_out=None if _post.punch_out==None else _post.punch_out.strftime("%Y-%m-%d %H:%M:%S")
                if True:#_post.group_no!=0 and _post.group_no!="0":
                    post_data = {  # 按照api格式塞進去
                        'attend_no':_post.attend_no,
                        'attend_date':_post.attend_date,
                        'comp_name':_post.comp_name,
                        'shift_no':_post.shift_no,
                        'emp_no': _post.emp_no,
                        'emp_name':_post.emp_name,
                        'group_no':_post.group_no,
                        'punch_in':punch_in,
                        'punch_out':punch_out,
                        'in_position':_post.in_position,
                        'out_position':_post.out_position,
                        'overtime_no':_post.overtime_no,
                        'salary_no':_post.salary_no,
                        'total_time':_post.total_time,
                        'fulltime':_post.fulltime,
                        'latetime':_post.latetime,
                        'attend_id':_post.attend_id,
                        'tel':_post.tel,
                        'work_position':_post.work_position,
                        'in_addr':_post.in_addr,
                        'out_arrd':_post.out_arrr
                    }
                    posts.append(post_data)

            if posts == []:
                Response = {
                    'status': 'Null',
                    'data': posts
                }
            else:
                Response = {
                    'status': 'Succeed',
                    'data': posts
                }
            #return Response
        except Exception as e:
            Response = {
                'status': 'Error',
                'msg': str(e)
            }
    elif  emp_no!= None and year!=None and month!=None: 
        try:
            all_posts = AttendanceEmpView.objects.filter(Q(emp_no=request.GET.get('emp_no')) & Q(attend_date__year=request.GET.get('year')) & Q(attend_date__month=request.GET.get('month')) & Q(working=1))  # 自模型取出資料
            posts = []

            for _post in all_posts:  # 按照api格式塞進去
                punch_in=None if _post.punch_in==None else _post.punch_in.strftime("%Y-%m-%d %H:%M:%S")#三元運算子
                punch_out=None if _post.punch_out==None else _post.punch_out.strftime("%Y-%m-%d %H:%M:%S")
                if True:#_post.group_no!=0 and _post.group_no!="0":
                    post_data = {  # 按照api格式塞進去
                        'attend_no':_post.attend_no,
                        'attend_date':_post.attend_date,
                        'comp_name':_post.comp_name,
                        'shift_no':_post.shift_no,
                        'emp_no': _post.emp_no,
                        'emp_name':_post.emp_name,
                        'group_no':_post.group_no,
                        'punch_in':punch_in,
                        'punch_out':punch_out,
                        'in_position':_post.in_position,
                        'out_position':_post.out_position,
                        'overtime_no':_post.overtime_no,
                        'salary_no':_post.salary_no,
                        'total_time':_post.total_time,
                        'fulltime':_post.fulltime,
                        'latetime':_post.latetime,
                        'attend_id':_post.attend_id,
                        'tel':_post.tel,
                        'work_position':_post.work_position
                    }
                    posts.append(post_data)

            if posts == []:
                Response = {
                    'status': 'Null',
                    'data': posts
                }
            else:
                Response = {
                    'status': 'Succeed',
                    'data': posts
                }
            #return Response
        except Exception as e:
            Response = {
                'status': 'Error',
                'msg': str(e)
            }
    elif  year!=None and month!=None and day!=None: 
        try:
            all_posts = AttendanceEmpView.objects.filter(Q(attend_date__year=request.GET.get('year')) & Q(attend_date__month=request.GET.get('month')) & Q(attend_date__day=request.GET.get('day')) & Q(working=1) )  # 自模型取出資料
            posts = []
            for _post in all_posts:  # 按照api格式塞進去
                punch_in=None if _post.punch_in==None else _post.punch_in.strftime("%Y-%m-%d %H:%M:%S")#三元運算子
                punch_out=None if _post.punch_out==None else _post.punch_out.strftime("%Y-%m-%d %H:%M:%S")
                if True:#_post.group_no!=0 and _post.group_no!="0":
                    post_data = {  # 按照api格式塞進去
                        'attend_no':_post.attend_no,
                        'attend_date':_post.attend_date,
                        'comp_name':_post.comp_name,
                        'shift_no':_post.shift_no,
                        'emp_no': _post.emp_no,
                        'emp_name':_post.emp_name,
                        'group_no':_post.group_no,
                        'punch_in':punch_in,
                        'punch_out':punch_out,
                        'in_position':_post.in_position,
                        'out_position':_post.out_position,
                        'overtime_no':_post.overtime_no,
                        'salary_no':_post.salary_no,
                        'total_time':_post.total_time,
                        'fulltime':_post.fulltime,
                        'latetime':_post.latetime,
                        'attend_id':_post.attend_id,
                        'tel':_post.tel,
                        'work_position':_post.work_position

                    }
                    posts.append(post_data)

            if posts == []:
                Response = {
                    'status': 'Null',
                    'data': posts
                }
            else:
                Response = {
                    'status': 'Succeed',
                    'data': posts
                }
            #return Response
        except Exception as e:
            Response = {
                'status': 'Error',
                'msg': str(e)
            }     
    elif  year!=None and month!=None: 
        try:
            all_posts = AttendanceEmpView.objects.filter(Q(attend_date__year=request.GET.get('year')) & Q(attend_date__month=request.GET.get('month')) & Q(working=1)  )  # 自模型取出資料
            posts = []

            for _post in all_posts:  # 按照api格式塞進去
                punch_in=None if _post.punch_in==None else _post.punch_in.strftime("%Y-%m-%d %H:%M:%S")#三元運算子
                punch_out=None if _post.punch_out==None else _post.punch_out.strftime("%Y-%m-%d %H:%M:%S")
                if True:#_post.group_no!=0 and _post.group_no!="0":
                    post_data = {  # 按照api格式塞進去
                        'attend_no':_post.attend_no,
                        'attend_date':_post.attend_date,
                        'comp_name':_post.comp_name,                    
                        'shift_no':_post.shift_no,
                        'emp_no': _post.emp_no,
                        'emp_name':_post.emp_name,
                        'group_no':_post.group_no,
                        'punch_in':punch_in,
                        'punch_out':punch_out,
                        'in_position':_post.in_position,
                        'out_position':_post.out_position,
                        'overtime_no':_post.overtime_no,
                        'salary_no':_post.salary_no,
                        'total_time':_post.total_time,
                        'fulltime':_post.fulltime,
                        'latetime':_post.latetime,
                        'attend_id':_post.attend_id,
                        'tel':_post.tel,
                        'work_position':_post.work_position

                    }
                    posts.append(post_data)

            if posts == []:
                Response = {
                    'status': 'Null',
                    'data': posts
                }
            else:
                Response = {
                    'status': 'Succeed',
                    'data': posts
                }
            #return Response
        except Exception as e:
            Response = {
                'status': 'Error',
                'msg': str(e)
            }
    else:
        try:
            all_posts = AttendanceEmpTodayView.objects.filter(working=1)
            posts = []

            for _post in all_posts:  # 按照api格式塞進去
                punch_in=None if _post.punch_in==None else _post.punch_in.strftime("%Y-%m-%d %H:%M:%S")#三元運算子
                punch_out=None if _post.punch_out==None else _post.punch_out.strftime("%Y-%m-%d %H:%M:%S")
                if True:#_post.group_no!=0 and _post.group_no!="0":
                    post_data = {  # 按照api格式塞進去
                        'attend_no':_post.attend_no,
                        'attend_date':_post.attend_date,
                        'comp_name':_post.comp_name,
                        'shift_no':_post.shift_no,
                        'emp_no': _post.emp_no,
                        'emp_name':_post.emp_name,
                        'group_no':_post.group_no,
                        'punch_in':punch_in,
                        'punch_out':punch_out,
                        'in_position':_post.in_position,
                        'out_position':_post.out_position,
                        'overtime_no':_post.overtime_no,
                        'salary_no':_post.salary_no,
                        'total_time':_post.total_time,
                        'fulltime':_post.fulltime,
                        'latetime':_post.latetime,
                        'attend_id':_post.attend_id,
                        'tel':_post.tel,
                        'work_position':_post.work_position

                    }
                    posts.append(post_data)

            if posts == []:
                Response = {
                    'status': 'Null',
                    'data': posts
                }
            else:
                Response = {
                    'status': 'Succeed',
                    'data': posts
                }
            #return Response
        except Exception as e:
            Response = {
                'status': 'Error',
                'msg': str(e)
            }

    
    return Response

def _get_time_period(request):
    try:
        jwt_verify=_verify_jwt(request.headers['Token'])
        if "list index out of range" in str(jwt_verify):
            raise Exception("token不存在!")
        elif "over limit!" in str(jwt_verify):
            raise Exception("token已過期!")
        all_posts = TimePeriod.objects.all()
        posts = []

        for _post in all_posts:  # 按照api格式塞進去
            post_data = {  # 按照api格式塞進去
            'id':_post.id,
            'startTime':_post.starttime.strftime("%H:%M:%S"),
            'endTime':_post.endtime.strftime("%H:%M:%S"),
            'crossDay':_post.crossday
             }
            posts.append(post_data)

            if posts == []:
                Response = {
                    'status': 'Null',
                    'data': posts
                }
            else:
                Response = {
                    'status': 'Succeed',
                    'data': posts
                }
            #return Response
    except Exception as e:
        Response = {
            'status': 'Error',
            'msg': str(e)
        }

    
    return Response

def _insert_time_period(request):
    try:
        jwt_verify=_verify_jwt(request.headers['Token'])
        if "list index out of range" in str(jwt_verify):
            raise Exception("token不存在!")
        elif "over limit!" in str(jwt_verify):
            raise Exception("token已過期!")
    except Exception as e:
        Response=str(e)
        return Response
    requestBody = request.body
    requestData = json.loads(requestBody)
    errorMsg = 'ERROR!'

    try:
        newTimePeriod = TimePeriod.objects.create(
            id=requestData['id'],
            starttime=requestData['startTime'],
            endtime=requestData['endTime'],
            crossday=requestData['crossDay']
        )
    except Exception as e:
        if 'Duplicate entry' in str(e):
            errorMsg += '該編號重複，請再確認編號'
        else:
            errorMsg += str(e)
        return errorMsg

    return 'Succeed'

def _update_time_period(request):  # 更新出勤資料
    try:
        jwt_verify=_verify_jwt(request.headers['Token'])
        if "list index out of range" in str(jwt_verify):
            raise Exception("token不存在!")
        elif "over limit!" in str(jwt_verify):
            raise Exception("token已過期!")
    except Exception as e:
        Response = {
                'status': 'Error',
                'msg': str(e)
            }
        return Response

    requestBody = request.body
    requestData = json.loads(requestBody)
    errorMsg = '修改錯誤!'
    try:
        _post = TimePeriod.objects.filter(id=requestData['id'])
        if _post[0].id:
            _post.update(
                starttime=requestData['startTime'],
                endtime=requestData['endTime'],
                crossday=requestData['crossDay']
            )

    except Exception as e:
        if "list index out of range" in str(e):
            errorMsg += "id不存在!"
        else: 
            errorMsg += str(e)
        return errorMsg

    return 'Succeed'

def _get_WorkpositionList(request):
    try:
        jwt_verify=_verify_jwt(request.headers['Token'])
        if "list index out of range" in str(jwt_verify):
            raise Exception("token不存在!")
        elif "over limit!" in str(jwt_verify):
            raise Exception("token已過期!")
        all_posts = WorkpositionList.objects.all()
        posts = []

        for _post in all_posts:  # 按照api格式塞進去
            post_data = {  # 按照api格式塞進去
            'wp_name':_post.wp_name,
            'wp_code':_post.wp_code,
             }
            posts.append(post_data)

            if posts == []:
                Response = {
                    'status': 'Null',
                    'data': posts
                }
            else:
                Response = {
                    'status': 'Succeed',
                    'data': posts
                }
            #return Response
    except Exception as e:
        Response = {
            'status': 'Error',
            'msg': str(e)
        }

    
    return Response

def _delete_attend_punch(request):
    try:
        jwt_verify=_verify_jwt(request.headers['Token'])
        if "list index out of range" in str(jwt_verify):
            raise Exception("token不存在!")
        elif "over limit!" in str(jwt_verify):
            raise Exception("token已過期!")
    except Exception as e:
        Response=str(e)
        return Response
    requestBody = request.body
    requestData = json.loads(requestBody)
    errorMsg = 'ERROR!'
    attend_no=requestData['attend_no']
    try:
        _post=AttendanceTable.objects.filter(attend_no=attend_no)
        if _post[0].punch_in != None and  _post[0].punch_out != None:
            _post.update(
                punch_out=None,
                out_position=None
            )
            return 'Update Succeed'
        elif _post[0].punch_in != None and  _post[0].punch_out == None:
            _post[0].delete()
            _post_shift=ShiftsTable.objects.filter(attend_no=attend_no)
            _post_shift.update(
                attend_no=None
            )
            return 'Delete Succeed'
    except Exception as e:
        if 'list index out of range' in str(e):
            errorMsg += '找不到此編號!請再確認編號!'
        else:
            errorMsg += str(e)
        return errorMsg

